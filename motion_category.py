# Ubuntu Version
import time
import undetected_chromedriver as webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

class motion_scraper:
    def __init__(self):
        self.driver = None
        self.products = []
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove the default sheet immediately
        
    def run_browser(self, url):
        service = Service(executable_path=r'Home/Downloads/chromedriver-mac-x64/chromedriver')
        chrome_options = webdriver.ChromeOptions()
        #chrome_options.add_argument("--headless")
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--no-sandbox')
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        self.driver.maximize_window()
        self.driver.get(url)
        
    def load_all_items(self):
        driver = self.driver
        unique_products = set()  # To track unique product names
        last_len = 0  # To track the length of unique_products after the last click
        
        while True:
            
            # Allow time for new items to load. Adjust sleep time as necessary.
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, './/div[contains(@class,"link-wrap")]//div[contains(@class, "text-wrapper")]')))
            
            # Retrieve all product names currently loaded
            product_names = driver.find_elements(By.XPATH, './/div[@class = "text"]')
            print(len(product_names))
            # Add product names to the set of unique products
            new_items_found = False
            for product in product_names:
                product_text = product.text.strip()
                if product_text and product_text not in unique_products:
                    unique_products.add(product_text)
                    new_items_found = True
            try:
                # Find and click the "Next" button
                next_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '//button[contains(@class, "btn-next") ]'))
                )
                driver.execute_script("arguments[0].click();", next_button)
            except:
                print ("Next button is not existant")
                break
            # Break the loop if no new items were found during this iteration
            if not new_items_found and len(unique_products) == last_len:
                break
            
            last_len = len(unique_products)  # Update last_len for the next iteration

        # Return the unique product names
        return list(unique_products)
    
    def scrape_website(self):
        url = "https://www.motion.com/"
        self.run_browser(url)
        try:
            accept_btn = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//button[contains(@id,"onetrust-accept-btn-handler")]')))
            accept_btn.click()
            
        except:
            print("error")
        time.sleep(5)
        categories_count = len(WebDriverWait(self.driver, 30).until(EC.presence_of_all_elements_located((By.XPATH, './/div[(contains(@class, "taxonomy-list"))]//li[(@class = "ng-star-inserted")]/a'))))
        print(categories_count)
        
        for index in range(categories_count):
            category_lis = WebDriverWait(self.driver, 30).until(EC.presence_of_all_elements_located((By.XPATH, './/div[(contains(@class, "taxonomy-list"))]//li[(@class = "ng-star-inserted")]/a')))
            category_li = category_lis[index]
            
            category_name = category_li.text.replace('/', '&')
            print(category_name)
            
        
    
        
            if category_name in self.workbook.sheetnames:
                sheet = self.workbook[category_name]
            else:
                sheet = self.workbook.create_sheet(title=category_name)
                sheet.append(["Sub Categories"])  # Add "Sub Categories" header to each new sheet
            
            try:
                
                category_li.click()
                
                # Assuming you've already navigated to the target page and initialized the driver
                sub_categories_texts = self.load_all_items()

                # Print and write the product names to the worksheet
                for idx, name in enumerate(sub_categories_texts, start=1):
                    print(name)
                    sheet.append([name])
                
                
                self.driver.back()

                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, './/a[contains(@class, "ng-star-inserted")]')))

            except Exception as e:
                print(f"Failed to navigate and scrape category: {e}")
                self.driver.back()
            
            
        time.sleep(1)
    
    def save_workbook(self, filename):
        self.workbook.save(filename=filename)
        
    

if __name__ == "__main__":
    scraper = motion_scraper()
    scraper.scrape_website()
    scraper.save_workbook('motion.com_sub_categories.xlsx')
